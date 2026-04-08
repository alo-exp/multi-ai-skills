"""
Unit tests for skills/comparator/matrix_builder.py

UT-MB-01: build_matrix creates valid XLSX
UT-MB-02: Title row is merged and correct
UT-MB-03: Header row has correct columns
UT-MB-04: Platform columns match input
UT-MB-05: Feature rows have correct ticks
UT-MB-06: Category rows are merged headings
UT-MB-07: Score and COUNTIF rows contain formulas
UT-MB-08: Return value has correct counts
UT-MB-09: Empty platforms list produces minimal matrix
"""

import json
import sys
import tempfile
from pathlib import Path

import openpyxl

COMPARATOR_DIR = str(Path(__file__).parent.parent / "skills" / "comparator")
if COMPARATOR_DIR not in sys.path:
    sys.path.insert(0, COMPARATOR_DIR)

import unittest.mock

from matrix_builder import build_matrix, main, TICK, DATA_START, HEADER_ROW, TITLE_ROW

# ── Test config ────────────────────────────────────────────────────────────────

SAMPLE_CONFIG = {
    "title": "Test Matrix",
    "categories": [
        {
            "name": "1. Deployment",
            "features": [
                {"name": "GitOps", "priority": "High"},
                {"name": "Canary", "priority": "Medium"},
            ],
        },
        {
            "name": "2. Monitoring",
            "features": [
                {"name": "Metrics", "priority": "Critical"},
            ],
        },
    ],
    "platforms": [
        {"name": "PlatformA", "features": ["GitOps", "Canary", "Metrics"]},
        {"name": "PlatformB", "features": ["GitOps"]},
    ],
}


def _build(config=None):
    """Build a matrix from config, return (result_dict, Workbook)."""
    cfg = config or SAMPLE_CONFIG
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = f.name
    result = build_matrix(cfg, out)
    wb = openpyxl.load_workbook(out)
    return result, wb


class TestMatrixBuilderBasic:
    def test_ut_mb_01_creates_valid_xlsx(self):
        """UT-MB-01: build_matrix produces a loadable XLSX."""
        result, wb = _build()
        assert wb.active is not None
        assert result["output"].endswith(".xlsx")

    def test_ut_mb_02_title_row(self):
        """UT-MB-02: Row 1 has the merged title."""
        _, wb = _build()
        ws = wb.active
        assert ws.cell(TITLE_ROW, 1).value == "Test Matrix"

    def test_ut_mb_03_header_row(self):
        """UT-MB-03: Row 2 has Capability/Feature, Priority, and platform names."""
        _, wb = _build()
        ws = wb.active
        assert ws.cell(HEADER_ROW, 1).value == "Capability / Feature"
        assert ws.cell(HEADER_ROW, 2).value == "Priority"
        assert ws.cell(HEADER_ROW, 3).value == "PlatformA"
        assert ws.cell(HEADER_ROW, 4).value == "PlatformB"

    def test_ut_mb_04_platform_columns(self):
        """UT-MB-04: Number of platform columns matches input."""
        result, _ = _build()
        assert result["platforms_added"] == 2
        assert result["platform_names"] == ["PlatformA", "PlatformB"]


class TestMatrixBuilderContent:
    def test_ut_mb_05_feature_ticks(self):
        """UT-MB-05: PlatformA has all 3 ticks; PlatformB has 1."""
        _, wb = _build()
        ws = wb.active
        # First feature row is DATA_START + 1 (after category heading)
        gitops_row = DATA_START + 1
        assert ws.cell(gitops_row, 3).value == TICK  # PlatformA: GitOps
        assert ws.cell(gitops_row, 4).value == TICK  # PlatformB: GitOps

        canary_row = DATA_START + 2
        assert ws.cell(canary_row, 3).value == TICK  # PlatformA: Canary
        assert ws.cell(canary_row, 4).value is None  # PlatformB: no Canary

    def test_ut_mb_06_category_rows(self):
        """UT-MB-06: Category rows contain heading text."""
        _, wb = _build()
        ws = wb.active
        assert ws.cell(DATA_START, 1).value == "1. Deployment"
        # Second category: row = DATA_START + 3 (cat heading + 2 features)
        assert ws.cell(DATA_START + 3, 1).value == "2. Monitoring"

    def test_ut_mb_07_formula_rows(self):
        """UT-MB-07: Row 3 (COUNTIF) and Row 4 (Score) contain formulas."""
        _, wb = _build()
        ws = wb.active
        countif_val = ws.cell(3, 3).value  # COUNTIF for PlatformA
        score_val = ws.cell(4, 3).value  # Score for PlatformA
        assert countif_val is not None and "COUNTIF" in str(countif_val)
        assert score_val is not None and "COUNTIFS" in str(score_val)


class TestMatrixBuilderEdgeCases:
    def test_ut_mb_08_return_counts(self):
        """UT-MB-08: Return dict has correct feature and category counts."""
        result, _ = _build()
        assert result["categories"] == 2
        assert result["features"] == 3  # GitOps + Canary + Metrics

    def test_ut_mb_09_empty_platforms(self):
        """UT-MB-09: Zero platforms produces a minimal matrix."""
        config = {
            "title": "Empty",
            "categories": [{"name": "Cat", "features": [{"name": "F", "priority": "Low"}]}],
            "platforms": [],
        }
        result, wb = _build(config)
        assert result["platforms_added"] == 0
        assert result["features"] == 1


class TestMatrixBuilderCloneXlsx:
    def test_build_matrix_with_clone_xlsx_warns(self, caplog):
        """UT-MB-10: build_matrix logs warning when clone_xlsx is provided (line 176)."""
        import logging
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        with caplog.at_level(logging.WARNING, logger="matrix_builder"):
            result = build_matrix(SAMPLE_CONFIG, out, clone_xlsx="dummy.xlsx")
        assert "clone_xlsx is not yet implemented" in caplog.text
        assert result["platforms_added"] == 2


class TestMatrixBuilderCLI:
    def test_main_cli(self, tmp_path):
        """UT-MB-11: main() CLI builds matrix from --config and --out args (lines 339-352)."""
        config_file = tmp_path / "config.json"
        config_file.write_text(json.dumps(SAMPLE_CONFIG))
        out_file = tmp_path / "output.xlsx"
        with unittest.mock.patch("sys.argv", [
            "matrix_builder",
            "--config", str(config_file),
            "--out", str(out_file),
        ]):
            main()
        assert out_file.exists()
        wb = openpyxl.load_workbook(str(out_file))
        assert wb.active.cell(1, 1).value == "Test Matrix"

    def test_main_cli_with_clone_style(self, tmp_path):
        """UT-MB-12: main() CLI with --clone-style arg still builds the matrix."""
        config_file = tmp_path / "config.json"
        config_file.write_text(json.dumps(SAMPLE_CONFIG))
        out_file = tmp_path / "output.xlsx"
        with unittest.mock.patch("sys.argv", [
            "matrix_builder",
            "--config", str(config_file),
            "--out", str(out_file),
            "--clone-style", "dummy.xlsx",
        ]):
            main()
        assert out_file.exists()
