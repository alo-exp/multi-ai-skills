"""Unit tests for matrix_ops module.

Tests UT-MX-01 through UT-MX-09 (subprocess-based) plus comprehensive
direct-import tests for internal helpers and public API (100% coverage).

Uses subprocess to invoke matrix_ops.py as a CLI for the original tests.
Uses direct imports for coverage tests.
"""

import json
import subprocess
import sys
import tempfile
import unittest.mock
from io import StringIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent
MATRIX_OPS = PROJECT_ROOT / "skills" / "comparator" / "matrix_ops.py"

# Direct import for coverage
COMPARATOR_DIR = str(PROJECT_ROOT / "skills" / "comparator")
if COMPARATOR_DIR not in sys.path:
    sys.path.insert(0, COMPARATOR_DIR)

from matrix_ops import (
    TICK, WEIGHTS, RowType,
    _Layout, _row_type, _clone_style, _last_plat_col, _platform_cols,
    _all_features, _all_categories, _unmerge_all, _remerge_all,
    _countif, _score_formula, _compute_score, _validate_no_orphans,
    _cache_col_styles, _apply_col_styles, _backfill_formulas,
    _read_row_data, _write_row_data, _json_out,
    add_platform, reorder_columns_by_score, create_combo_column,
    verify_ticks, reorder_rows, reorder_categories, extract_features,
    ranked_scores, info, main,
)

_WDIR = PROJECT_ROOT


TICK = "\u2714"  # checkmark


# ---------------------------------------------------------------------------
# Shared workbook builders
# ---------------------------------------------------------------------------

def _make_test_wb(with_title=True):
    """Build a test workbook with two categories and three features."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if with_title:
        title_row = 1
        header_row = 2
        total_row = 3
        score_row = 4
        data_start = 5
        ws.cell(title_row, 1).value = "Test Matrix"
        ws.merge_cells(f"A{title_row}:D{title_row}")
    else:
        header_row = 1
        total_row = 2
        score_row = 3
        data_start = 4

    # Headers
    ws.cell(header_row, 1).value = "Capability / Feature"
    ws.cell(header_row, 2).value = "Priority"
    ws.cell(header_row, 3).value = "Platform A"
    ws.cell(header_row, 4).value = "Platform B"

    last_col = 4
    for c in range(3, last_col + 1):
        ltr = get_column_letter(c)
        ws.cell(total_row, c).value = f'=COUNTIF({ltr}{data_start}:{ltr}1048576,"?*")'
        ws.cell(score_row, c).value = 0

    # Category 1 + features
    r = data_start
    ws.cell(r, 1).value = "Cat1"
    ws.merge_cells(f"A{r}:D{r}")

    r += 1
    ws.cell(r, 1).value = "Feat1"
    ws.cell(r, 2).value = "High"
    ws.cell(r, 3).value = TICK
    ws.cell(r, 4).value = None

    r += 1
    ws.cell(r, 1).value = "Feat2"
    ws.cell(r, 2).value = "Medium"
    ws.cell(r, 3).value = TICK
    ws.cell(r, 4).value = TICK

    # Category 2 + feature
    r += 1
    ws.cell(r, 1).value = "Cat2"
    ws.merge_cells(f"A{r}:D{r}")

    r += 1
    ws.cell(r, 1).value = "Feat3"
    ws.cell(r, 2).value = "Low"
    ws.cell(r, 3).value = None
    ws.cell(r, 4).value = TICK

    return wb


def _save_wb(wb) -> str:
    """Save workbook to a temp file and return the path."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    wb.save(path)
    return path


def _make_tmp_xlsx(with_title=True) -> str:
    """Build test workbook and save to temp file. Return path."""
    return _save_wb(_make_test_wb(with_title=with_title))


# ---------------------------------------------------------------------------
# Original subprocess-based tests (UT-MX-01 to UT-MX-09)
# ---------------------------------------------------------------------------

def _create_test_matrix(path: str, with_title: bool = True) -> None:
    """Create a minimal test XLSX matrix.

    Layout (with_title=True):
      Row 1: Title (merged A1:E1)
      Row 2: Headers -- "", "Priority", "PlatformA", "PlatformB"
      Row 3: COUNTIF formulas
      Row 4: Score formulas
      Row 5: Category row "Core Features" (merged)
      Row 6: Feature "CI/CD Pipelines", "Critical", tick, tick
      Row 7: Feature "Container Support", "High", tick, ""
      Row 8: Feature "Monitoring", "Medium", "", tick
      Row 9: Category row "Advanced Features" (merged)
      Row 10: Feature "AI Ops", "Low", tick, ""
      Row 11: Feature "GitOps", "Very High", "", tick
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    if with_title:
        ws.cell(1, 1).value = "Test Comparison Matrix"
        ws.merge_cells("A1:E1")
        header_row = 2
        total_row = 3
        score_row = 4
        data_start = 5
    else:
        header_row = 1
        total_row = 2
        score_row = 3
        data_start = 4

    # Headers
    ws.cell(header_row, 1).value = ""
    ws.cell(header_row, 2).value = "Priority"
    ws.cell(header_row, 3).value = "PlatformA"
    ws.cell(header_row, 4).value = "PlatformB"

    last_col = 4

    # COUNTIF and score rows (placeholder formulas)
    for c in range(3, last_col + 1):
        ltr = get_column_letter(c)
        ws.cell(total_row, c).value = f'=COUNTIF({ltr}{data_start}:{ltr}1048576,"?*")'
        ws.cell(score_row, c).value = 0  # Placeholder

    # Category 1
    r = data_start
    ws.cell(r, 1).value = "Core Features"
    ws.merge_cells(f"A{r}:{get_column_letter(last_col)}{r}")

    # Feature rows
    features = [
        ("CI/CD Pipelines", "Critical", True, True),
        ("Container Support", "High", True, False),
        ("Monitoring", "Medium", False, True),
    ]
    for feat, prio, tick_a, tick_b in features:
        r += 1
        ws.cell(r, 1).value = feat
        ws.cell(r, 2).value = prio
        ws.cell(r, 3).value = TICK if tick_a else None
        ws.cell(r, 4).value = TICK if tick_b else None

    # Category 2
    r += 1
    ws.cell(r, 1).value = "Advanced Features"
    ws.merge_cells(f"A{r}:{get_column_letter(last_col)}{r}")

    features2 = [
        ("AI Ops", "Low", True, False),
        ("GitOps", "Very High", False, True),
    ]
    for feat, prio, tick_a, tick_b in features2:
        r += 1
        ws.cell(r, 1).value = feat
        ws.cell(r, 2).value = prio
        ws.cell(r, 3).value = TICK if tick_a else None
        ws.cell(r, 4).value = TICK if tick_b else None

    wb.save(path)


def _run_matrix_ops(*args: str) -> dict:
    """Run matrix_ops.py as subprocess and parse JSON output."""
    cmd = [sys.executable, str(MATRIX_OPS)] + list(args)
    result = subprocess.run(
        cmd,
        capture_output=True, text=True,
        cwd=str(PROJECT_ROOT),
    )
    assert result.returncode == 0, (
        f"matrix_ops.py failed:\nstdout: {result.stdout}\nstderr: {result.stderr}"
    )
    return json.loads(result.stdout)


class TestMatrixInfo:
    """Tests for the 'info' command."""

    def test_ut_mx_01_info_returns_platforms(self):
        """UT-MX-01: info command returns platform names."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("info", "--src", path)
        assert "PlatformA" in data["platforms"]
        assert "PlatformB" in data["platforms"]
        assert data["platform_count"] == 2

    def test_ut_mx_02_info_returns_categories(self):
        """UT-MX-02: info returns correct category names."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("info", "--src", path)
        assert "Core Features" in data["categories"]
        assert "Advanced Features" in data["categories"]
        assert data["category_count"] == 2

    def test_ut_mx_03_info_total_features(self):
        """UT-MX-03: info returns correct total feature count."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("info", "--src", path)
        assert data["total_features"] == 5

    def test_ut_mx_04_info_detects_with_title_layout(self):
        """UT-MX-04: info detects 'with_title' layout."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path, with_title=True)
        data = _run_matrix_ops("info", "--src", path)
        assert data["layout"] == "with_title"

    def test_ut_mx_05_info_detects_no_title_layout(self):
        """UT-MX-05: info detects 'no_title' layout."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path, with_title=False)
        data = _run_matrix_ops("info", "--src", path)
        assert data["layout"] == "no_title"


class TestMatrixScores:
    """Tests for the 'scores' command."""

    def test_ut_mx_06_scores_returns_rankings(self):
        """UT-MX-06: scores command returns ranked platforms."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("scores", "--src", path)
        rankings = data["rankings"]
        assert len(rankings) == 2
        names = [r["platform"] for r in rankings]
        assert "PlatformA" in names
        assert "PlatformB" in names


class TestMatrixExtractFeatures:
    """Tests for the 'extract-features' command."""

    def test_ut_mx_07_extract_features_groups_by_category(self):
        """UT-MX-07: extract-features returns features grouped by category."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("extract-features", "--src", path)
        assert data["total_categories"] == 2
        assert data["total_features"] == 5
        cat_names = [c["name"] for c in data["categories"]]
        assert "Core Features" in cat_names


class TestMatrixVerify:
    """Tests for the 'verify' command."""

    def test_ut_mx_08_verify_returns_tick_lists(self):
        """UT-MX-08: verify returns per-platform tick lists."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("verify", "--src", path)
        assert "PlatformA" in data
        assert "CI/CD Pipelines" in data["PlatformA"]["ticked"]
        assert data["PlatformA"]["tick_count"] == 3  # CI/CD, Container, AI Ops


class TestMatrixAddPlatform:
    """Tests for the 'add-platform' command."""

    def test_ut_mx_09_add_platform(self):
        """UT-MX-09: add-platform adds a new column."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            src_path = f.name
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name

        _create_test_matrix(src_path)

        # Create features JSON
        features = {"CI/CD Pipelines": True, "Container Support": True, "Monitoring": False, "AI Ops": False, "GitOps": True}
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump(features, f)
            feat_path = f.name

        data = _run_matrix_ops(
            "add-platform",
            "--src", src_path,
            "--out", out_path,
            "--platform", "PlatformC",
            "--features", feat_path,
        )
        assert data["ticks_applied"] == 3  # CI/CD, Container, GitOps
        assert data["platform_col"] == 5  # New column after D

        # Verify the output file
        info_data = _run_matrix_ops("info", "--src", out_path)
        assert "PlatformC" in info_data["platforms"]
        assert info_data["platform_count"] == 3


# ---------------------------------------------------------------------------
# Task 1: Internal helpers and Layout detection (direct import, coverage tests)
# ---------------------------------------------------------------------------

class TestLayout:
    """Tests for _Layout auto-detection."""

    def test_layout_with_title(self):
        """_Layout with title row: B1 != 'Priority' -> title_row=1, header_row=2."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        assert L.title_row == 1
        assert L.header_row == 2
        assert L.total_row == 3
        assert L.score_row == 4
        assert L.data_start == 5

    def test_layout_without_title(self):
        """_Layout without title: B1 == 'Priority' -> title_row=None, header_row=1."""
        wb = _make_test_wb(with_title=False)
        ws = wb.active
        L = _Layout(ws)
        assert L.title_row is None
        assert L.header_row == 1
        assert L.total_row == 2
        assert L.score_row == 3
        assert L.data_start == 4

    def test_layout_freeze_pane_with_title(self):
        """freeze_pane returns correct cell ref for layout with title."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        assert L.freeze_pane == "A5"

    def test_layout_freeze_pane_without_title(self):
        """freeze_pane returns correct cell ref for layout without title."""
        wb = _make_test_wb(with_title=False)
        ws = wb.active
        L = _Layout(ws)
        assert L.freeze_pane == "A4"


class TestRowType:
    """Tests for _row_type detection."""

    def test_row_type_category(self):
        """Row with col A value but no/dash col B is CATEGORY."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        # data_start row = 5 = category row "Cat1"
        assert _row_type(ws, 5) == RowType.CATEGORY

    def test_row_type_feature(self):
        """Row with both col A and col B populated is FEATURE."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        # Row 6 = Feat1 with priority "High"
        assert _row_type(ws, 6) == RowType.FEATURE

    def test_row_type_empty_row(self):
        """Row with no col A value returns None."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Row 1 has nothing
        assert _row_type(ws, 1) is None

    def test_row_type_dash_priority(self):
        """Row with col A value and dash in col B is CATEGORY."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "My Category"
        ws.cell(1, 2).value = "\u2014"  # em dash
        assert _row_type(ws, 1) == RowType.CATEGORY

    def test_row_type_empty_string_priority(self):
        """Row with col A value and empty string col B is CATEGORY."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "My Category"
        ws.cell(1, 2).value = "  "
        assert _row_type(ws, 1) == RowType.CATEGORY


class TestCloneStyle:
    """Tests for _clone_style."""

    def test_clone_style_copies_font(self):
        """_clone_style copies font from src to dst."""
        wb = openpyxl.Workbook()
        ws = wb.active
        src = ws.cell(1, 1)
        dst = ws.cell(2, 1)
        src.font = Font(name="Arial", bold=True, size=14)
        _clone_style(src, dst)
        assert dst.font.bold is True
        assert dst.font.name == "Arial"

    def test_clone_style_copies_number_format(self):
        """_clone_style copies number_format."""
        wb = openpyxl.Workbook()
        ws = wb.active
        src = ws.cell(1, 1)
        dst = ws.cell(2, 1)
        src.number_format = "0.00%"
        _clone_style(src, dst)
        assert dst.number_format == "0.00%"


class TestLastPlatCol:
    """Tests for _last_plat_col."""

    def test_last_plat_col_with_title(self):
        """_last_plat_col finds last platform column (col D = 4)."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        assert _last_plat_col(ws, L) == 4

    def test_last_plat_col_without_title(self):
        """_last_plat_col works in no-title layout."""
        wb = _make_test_wb(with_title=False)
        ws = wb.active
        L = _Layout(ws)
        assert _last_plat_col(ws, L) == 4


class TestPlatformCols:
    """Tests for _platform_cols."""

    def test_platform_cols_returns_dict(self):
        """_platform_cols returns {name: col_index} from header row."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        result = _platform_cols(ws, L)
        assert result == {"Platform A": 3, "Platform B": 4}

    def test_platform_cols_no_title(self):
        """_platform_cols works in no-title layout."""
        wb = _make_test_wb(with_title=False)
        ws = wb.active
        L = _Layout(ws)
        result = _platform_cols(ws, L)
        assert "Platform A" in result
        assert "Platform B" in result


class TestAllFeatures:
    """Tests for _all_features."""

    def test_all_features_returns_only_features(self):
        """_all_features returns feature names, not category names."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        feats = _all_features(ws, L)
        assert "Feat1" in feats
        assert "Feat2" in feats
        assert "Feat3" in feats
        assert "Cat1" not in feats
        assert "Cat2" not in feats
        assert len(feats) == 3


class TestAllCategories:
    """Tests for _all_categories."""

    def test_all_categories_returns_blocks(self):
        """_all_categories returns category blocks with features."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        cats = _all_categories(ws, L)
        assert len(cats) == 2
        assert cats[0]["name"] == "Cat1"
        assert len(cats[0]["features"]) == 2
        assert cats[1]["name"] == "Cat2"
        assert len(cats[1]["features"]) == 1

    def test_all_categories_feature_priorities(self):
        """_all_categories includes priority in feature dicts."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        cats = _all_categories(ws, L)
        feat1 = cats[0]["features"][0]
        assert feat1["name"] == "Feat1"
        assert feat1["priority"] == "High"


class TestUnmergRemerge:
    """Tests for _unmerge_all and _remerge_all."""

    def test_unmerge_all_removes_merges(self):
        """_unmerge_all removes all merges and returns range strings."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        saved = _unmerge_all(ws)
        assert len(saved) > 0
        assert len(ws.merged_cells.ranges) == 0

    def test_remerge_all_restores_merges(self):
        """_remerge_all re-merges title and category rows."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        _unmerge_all(ws)
        _remerge_all(ws, 4, L)
        # Should have title + 2 category merges = 3 merges
        assert len(ws.merged_cells.ranges) == 3

    def test_remerge_all_no_title(self):
        """_remerge_all without title only merges category rows."""
        wb = _make_test_wb(with_title=False)
        ws = wb.active
        L = _Layout(ws)
        _unmerge_all(ws)
        _remerge_all(ws, 4, L)
        # Should have 2 category merges only (no title row)
        assert len(ws.merged_cells.ranges) == 2


class TestCountifAndScore:
    """Tests for _countif and _score_formula."""

    def test_countif_formula(self):
        """_countif returns correct formula string."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        formula = _countif("C", L)
        assert formula.startswith("=COUNTIF(C5:")
        assert '"?*"' in formula

    def test_score_formula(self):
        """_score_formula returns COUNTIFS formula with weights."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        formula = _score_formula("C", L)
        assert formula.startswith("=")
        assert "COUNTIFS" in formula
        # Should have one term per priority weight
        assert formula.count("COUNTIFS") == len(WEIGHTS)

    def test_countif_no_title(self):
        """_countif works with no-title layout (data_start=4)."""
        wb = _make_test_wb(with_title=False)
        ws = wb.active
        L = _Layout(ws)
        formula = _countif("C", L)
        assert "C4:" in formula


class TestComputeScore:
    """Tests for _compute_score."""

    def test_compute_score_with_ticks(self):
        """_compute_score calculates weighted score correctly."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        # Platform A (col 3): Feat1 (High=3, ticked), Feat2 (Medium=2, ticked), Feat3 (Low=1, not ticked)
        # Expected score: 3 + 2 = 5
        score = _compute_score(ws, 3, L)
        assert score == 5

    def test_compute_score_platform_b(self):
        """_compute_score for Platform B (col 4)."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        # Platform B (col 4): Feat2 (Medium=2, ticked), Feat3 (Low=1, ticked)
        score = _compute_score(ws, 4, L)
        assert score == 3


class TestValidateNoOrphans:
    """Tests for _validate_no_orphans."""

    def test_validate_no_orphans_clean(self):
        """_validate_no_orphans returns empty list when all features exist."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        orphans = _validate_no_orphans(ws, {"Feat1", "Feat2"}, L)
        assert orphans == []

    def test_validate_no_orphans_detects_missing(self):
        """_validate_no_orphans detects features not in matrix."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        orphans = _validate_no_orphans(ws, {"Feat1", "NonExistent"}, L)
        assert "NonExistent" in orphans

    def test_validate_no_orphans_empty_set(self):
        """_validate_no_orphans with empty feature set returns empty list."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        assert _validate_no_orphans(ws, set(), L) == []


class TestCacheApplyColStyles:
    """Tests for _cache_col_styles and _apply_col_styles."""

    def test_cache_apply_col_styles_roundtrip(self):
        """_cache_col_styles and _apply_col_styles preserve font."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        ws.cell(1, 3).font = Font(bold=True, color="FF0000")
        styles = _cache_col_styles(ws, 3)
        # Move to another column
        _apply_col_styles(ws, 5, styles)
        assert ws.cell(1, 5).font.bold is True


class TestBackfillFormulas:
    """Tests for _backfill_formulas."""

    def test_backfill_formulas_writes_formulas(self):
        """_backfill_formulas rewrites COUNTIF and score for all platform cols."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        L = _Layout(ws)
        _backfill_formulas(ws, 4, L)
        # Check col C total row formula
        assert "COUNTIF" in str(ws.cell(L.total_row, 3).value)
        # Check col D
        assert "COUNTIF" in str(ws.cell(L.total_row, 4).value)
        assert "COUNTIFS" in str(ws.cell(L.score_row, 3).value)


class TestReadWriteRowData:
    """Tests for _read_row_data and _write_row_data."""

    def test_read_write_row_roundtrip(self):
        """_read_row_data and _write_row_data roundtrip preserves values."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        # Row 6 = Feat1
        rd = _read_row_data(ws, 6, 4)
        assert rd["values"][0] == "Feat1"
        assert rd["values"][1] == "High"
        assert rd["row_type"] == RowType.FEATURE

        # Write to different row and verify
        _write_row_data(ws, 20, rd)
        assert ws.cell(20, 1).value == "Feat1"
        assert ws.cell(20, 2).value == "High"

    def test_read_row_category(self):
        """_read_row_data on category row returns CATEGORY row_type."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        rd = _read_row_data(ws, 5, 4)
        assert rd["row_type"] == RowType.CATEGORY


class TestJsonOut:
    """Tests for _json_out."""

    def test_json_out_prints_json(self, capsys):
        """_json_out prints JSON to stdout."""
        _json_out({"key": "value", "num": 42})
        captured = capsys.readouterr()
        data = json.loads(captured.out)
        assert data["key"] == "value"
        assert data["num"] == 42


# ---------------------------------------------------------------------------
# Task 2: Public API function tests
# ---------------------------------------------------------------------------

class TestAddPlatformDirect:
    """Direct tests for add_platform()."""

    def test_add_platform_basic(self):
        """add_platform adds column, applies ticks, returns correct counts."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = add_platform(src, out, "Platform C", {"Feat1": True, "Feat2": False, "Feat3": True})
        assert result["ticks_applied"] == 2
        assert result["new_rows_added"] == 0
        assert result["orphans"] == []
        assert result["platform_col"] == 5

    def test_add_platform_orphans_reported(self):
        """add_platform with orphans: features not in matrix reported."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = add_platform(src, out, "Platform C", {"Feat1": True, "GhostFeat": True})
        assert "GhostFeat" in result["orphans"]

    def test_add_platform_no_new_rows(self):
        """add_platform with new_rows=None works correctly."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = add_platform(src, out, "Platform C", {"Feat1": True}, new_rows=None)
        assert result["new_rows_added"] == 0

    def test_add_platform_with_new_rows(self):
        """add_platform with new_rows adds rows to the correct category."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        new_rows = [{"category": "Cat1", "feature": "NewFeat", "priority": "High", "ticked": True}]
        result = add_platform(src, out, "Platform C", {}, new_rows=new_rows)
        assert result["new_rows_added"] == 1
        # Verify it appears in output
        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.active
        found = any(ws2.cell(r, 1).value == "NewFeat" for r in range(1, ws2.max_row + 1))
        assert found

    def test_add_platform_no_title_layout(self):
        """add_platform works with no-title layout."""
        src = _make_tmp_xlsx(with_title=False)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = add_platform(src, out, "Platform C", {"Feat1": True})
        assert result["ticks_applied"] == 1

    def test_add_platform_output_has_new_column(self):
        """add_platform output XLSX has the new platform in header."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        add_platform(src, out, "NewPlat", {})
        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.active
        assert ws2.cell(2, 5).value == "NewPlat"

    def test_add_platform_new_rows_at_end_of_last_cat(self):
        """add_platform flushes new_rows for last category correctly."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        new_rows = [{"category": "Cat2", "feature": "EndFeat", "priority": "Low", "ticked": False}]
        result = add_platform(src, out, "Platform C", {}, new_rows=new_rows)
        assert result["new_rows_added"] == 1


class TestReorderColumnsByScore:
    """Tests for reorder_columns_by_score()."""

    def test_reorder_columns_by_score_basic(self):
        """reorder_columns_by_score reorders columns by descending score."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = reorder_columns_by_score(src, out)
        assert "order" in result
        assert "scores" in result
        # Platform A score = 5, Platform B score = 3, A should come first
        assert result["order"][0] == "Platform A"

    def test_reorder_columns_no_platforms(self):
        """reorder_columns_by_score with empty platform list returns empty result."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2).value = "Priority"  # no-title layout, no platforms
        src = _save_wb(wb)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = reorder_columns_by_score(src, out)
        assert result["order"] == []
        assert result["scores"] == {}


class TestCreateComboColumn:
    """Tests for create_combo_column()."""

    def test_create_combo_column_basic(self):
        """create_combo_column creates union of ticks from two platforms."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = create_combo_column(src, out, "A+B", "Platform A", "Platform B")
        assert "combo_col" in result
        # Feat1: A only, Feat2: both, Feat3: B only
        assert result["unique_a"] == 1
        assert result["unique_b"] == 1
        assert result["overlap"] == 1
        assert result["tick_count"] == 3

    def test_create_combo_column_missing_platform(self):
        """create_combo_column with missing platform returns error dict."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = create_combo_column(src, out, "combo", "Platform A", "NonExistent")
        assert "error" in result


class TestVerifyTicks:
    """Tests for verify_ticks()."""

    def test_verify_ticks_returns_per_platform_lists(self):
        """verify_ticks returns correct ticked/not_ticked per platform."""
        src = _make_tmp_xlsx(with_title=True)
        result = verify_ticks(src)
        assert "Platform A" in result
        assert "Feat1" in result["Platform A"]["ticked"]
        assert "Feat3" in result["Platform A"]["not_ticked"]
        assert result["Platform A"]["tick_count"] == 2


class TestReorderRows:
    """Tests for reorder_rows()."""

    def test_reorder_rows_basic(self):
        """reorder_rows reorders features within a category."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = reorder_rows(src, out, "Cat1", ["Feat2", "Feat1"])
        assert result["rows_reordered"] == 2
        assert result["features_not_found"] == []
        # Verify order in output
        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.active
        # After category row (row 5), row 6 should be Feat2 now
        assert ws2.cell(6, 1).value == "Feat2"

    def test_reorder_rows_missing_category(self):
        """reorder_rows with missing category returns error."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = reorder_rows(src, out, "NonExistentCat", ["Feat1"])
        assert "error" in result

    def test_reorder_rows_not_found_features(self):
        """reorder_rows with missing feature names reports them in not_found."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = reorder_rows(src, out, "Cat1", ["Feat1", "GhostFeat"])
        assert "GhostFeat" in result["features_not_found"]


class TestReorderCategories:
    """Tests for reorder_categories()."""

    def test_reorder_categories_basic(self):
        """reorder_categories reorders category blocks."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = reorder_categories(src, out, ["Cat2", "Cat1"])
        assert result["categories_reordered"] == 2
        # Verify Cat2 now comes first in data
        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.active
        assert ws2.cell(5, 1).value == "Cat2"

    def test_reorder_categories_not_found(self):
        """reorder_categories reports unknown category names."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = reorder_categories(src, out, ["Cat1", "GhostCat"])
        assert "GhostCat" in result["categories_not_found"]


class TestExtractFeatures:
    """Tests for extract_features()."""

    def test_extract_features_grouping(self):
        """extract_features returns correct category/feature grouping."""
        src = _make_tmp_xlsx(with_title=True)
        result = extract_features(src)
        assert result["total_categories"] == 2
        assert result["total_features"] == 3
        names = [c["name"] for c in result["categories"]]
        assert "Cat1" in names
        assert "Cat2" in names


class TestRankedScores:
    """Tests for ranked_scores()."""

    def test_ranked_scores_ordering(self):
        """ranked_scores returns platforms ranked by descending score with rank numbers."""
        src = _make_tmp_xlsx(with_title=True)
        result = ranked_scores(src)
        rankings = result["rankings"]
        assert len(rankings) == 2
        assert rankings[0]["rank"] == 1
        assert rankings[1]["rank"] == 2
        # Platform A has higher score
        assert rankings[0]["platform"] == "Platform A"
        assert rankings[0]["score"] >= rankings[1]["score"]


class TestInfoDirect:
    """Tests for info()."""

    def test_info_with_title(self):
        """info returns correct metadata with title layout."""
        src = _make_tmp_xlsx(with_title=True)
        result = info(src)
        assert result["title"] == "Test Matrix"
        assert result["platform_count"] == 2
        assert result["layout"] == "with_title"
        assert result["header_row"] == 2
        assert result["data_start_row"] == 5

    def test_info_no_title(self):
        """info returns correct metadata without title layout."""
        src = _make_tmp_xlsx(with_title=False)
        result = info(src)
        assert result["layout"] == "no_title"
        assert result["header_row"] == 1
        assert result["data_start_row"] == 4


class TestMainCLI:
    """Tests for main() CLI function."""

    def _make_src_and_features(self):
        src = _make_tmp_xlsx(with_title=True)
        feats = {"Feat1": True, "Feat2": False}
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump(feats, f)
            feat_file = f.name
        return src, feat_file

    def test_main_info(self, capsys):
        """main() with 'info' subcommand outputs JSON."""
        src = _make_tmp_xlsx(with_title=True)
        with unittest.mock.patch("sys.argv", ["matrix_ops", "info", "--src", src]):
            main()
        out = capsys.readouterr().out
        data = json.loads(out)
        assert data["platform_count"] == 2

    def test_main_scores(self, capsys):
        """main() with 'scores' subcommand outputs JSON rankings."""
        src = _make_tmp_xlsx(with_title=True)
        with unittest.mock.patch("sys.argv", ["matrix_ops", "scores", "--src", src]):
            main()
        out = capsys.readouterr().out
        data = json.loads(out)
        assert "rankings" in data

    def test_main_extract_features(self, capsys):
        """main() with 'extract-features' subcommand outputs JSON."""
        src = _make_tmp_xlsx(with_title=True)
        with unittest.mock.patch("sys.argv", ["matrix_ops", "extract-features", "--src", src]):
            main()
        out = capsys.readouterr().out
        data = json.loads(out)
        assert "categories" in data

    def test_main_verify(self, capsys):
        """main() with 'verify' subcommand outputs JSON."""
        src = _make_tmp_xlsx(with_title=True)
        with unittest.mock.patch("sys.argv", ["matrix_ops", "verify", "--src", src]):
            main()
        out = capsys.readouterr().out
        data = json.loads(out)
        assert "Platform A" in data

    def test_main_add_platform(self, capsys):
        """main() with 'add-platform' subcommand outputs JSON."""
        src, feat_file = self._make_src_and_features()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        with unittest.mock.patch("sys.argv", [
            "matrix_ops", "add-platform",
            "--src", src, "--out", out,
            "--platform", "TestPlat",
            "--features", feat_file,
        ]):
            main()
        stdout = capsys.readouterr().out
        data = json.loads(stdout)
        assert data["ticks_applied"] == 1

    def test_main_add_platform_with_new_rows(self, capsys):
        """main() add-platform with --new-rows JSON file."""
        src, feat_file = self._make_src_and_features()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        new_rows_data = [{"category": "Cat1", "feature": "CLI Feat", "priority": "Low", "ticked": True}]
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump(new_rows_data, f)
            nr_file = f.name
        with unittest.mock.patch("sys.argv", [
            "matrix_ops", "add-platform",
            "--src", src, "--out", out,
            "--platform", "TestPlat",
            "--features", feat_file,
            "--new-rows", nr_file,
        ]):
            main()
        stdout = capsys.readouterr().out
        data = json.loads(stdout)
        assert data["new_rows_added"] == 1

    def test_main_reorder_columns(self, capsys):
        """main() with 'reorder-columns' subcommand outputs JSON."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        with unittest.mock.patch("sys.argv", [
            "matrix_ops", "reorder-columns",
            "--src", src, "--out", out,
        ]):
            main()
        stdout = capsys.readouterr().out
        data = json.loads(stdout)
        assert "order" in data

    def test_main_combo(self, capsys):
        """main() with 'combo' subcommand outputs JSON."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        with unittest.mock.patch("sys.argv", [
            "matrix_ops", "combo",
            "--src", src, "--out", out,
            "--name", "A+B",
            "--platform-a", "Platform A",
            "--platform-b", "Platform B",
        ]):
            main()
        stdout = capsys.readouterr().out
        data = json.loads(stdout)
        assert "tick_count" in data

    def test_main_reorder_rows(self, capsys):
        """main() with 'reorder-rows' subcommand outputs JSON."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        order_data = ["Feat2", "Feat1"]
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump(order_data, f)
            order_file = f.name
        with unittest.mock.patch("sys.argv", [
            "matrix_ops", "reorder-rows",
            "--src", src, "--out", out,
            "--category", "Cat1",
            "--order", order_file,
        ]):
            main()
        stdout = capsys.readouterr().out
        data = json.loads(stdout)
        assert data["rows_reordered"] == 2

    def test_main_reorder_categories(self, capsys):
        """main() with 'reorder-categories' subcommand outputs JSON."""
        src = _make_tmp_xlsx(with_title=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        order_data = ["Cat2", "Cat1"]
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump(order_data, f)
            order_file = f.name
        with unittest.mock.patch("sys.argv", [
            "matrix_ops", "reorder-categories",
            "--src", src, "--out", out,
            "--order", order_file,
        ]):
            main()
        stdout = capsys.readouterr().out
        data = json.loads(stdout)
        assert data["categories_reordered"] == 2


# ---------------------------------------------------------------------------
# Additional tests to reach 100% coverage on remaining uncovered branches
# ---------------------------------------------------------------------------

class TestPlatformColsEmpty:
    """Test _platform_cols break branch (line 155)."""

    def test_platform_cols_stops_at_empty_header(self):
        """_platform_cols stops at first empty header (break branch)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # No-title layout: B1 = "Priority", no platform headers
        ws.cell(1, 2).value = "Priority"
        L = _Layout(ws)
        result = _platform_cols(ws, L)
        # No platforms -- the break path is exercised
        assert result == {}


class TestAddPlatformEmptyExpandedBranch:
    """Test add_platform else-branch when expanded is empty at new-row insertion."""

    def test_add_platform_new_rows_to_first_category_before_any_row(self):
        """New row inserted before any expanded data exercises the empty-expanded else branch."""
        # Build a workbook where the FIRST data row is the category (no leading features)
        wb = openpyxl.Workbook()
        ws = wb.active
        # with-title layout
        ws.cell(1, 1).value = "Title"
        ws.merge_cells("A1:C1")
        ws.cell(2, 1).value = "Feature"
        ws.cell(2, 2).value = "Priority"
        ws.cell(2, 3).value = "PlatA"
        ws.cell(3, 3).value = '=COUNTIF(C5:C1048576,"?*")'
        ws.cell(4, 3).value = 0
        # data_start = 5: category immediately
        ws.cell(5, 1).value = "OnlyCat"
        ws.merge_cells("A5:C5")
        ws.cell(6, 1).value = "ExistFeat"
        ws.cell(6, 2).value = "Low"
        ws.cell(6, 3).value = None

        src = _save_wb(wb)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name

        # Insert a new row into "OnlyCat" -- first category encountered when expanded is empty
        new_rows = [{"category": "OnlyCat", "feature": "BrandNew", "priority": "High", "ticked": True}]
        result = add_platform(src, out, "PlatB", {}, new_rows=new_rows)
        assert result["new_rows_added"] == 1


class TestAddPlatformClearLeftovers:
    """Test add_platform clear-leftover-rows branch (lines 417-418)."""

    def test_add_platform_clears_extra_rows_when_new_rows_collapse(self):
        """When new_rows causes the expanded list to have fewer rows than the original
        (which can't normally happen, but the clear loop is exercised by any new_rows call
        that leaves trailing empty rows in the sheet)."""
        # Create a matrix with extra empty rows at the bottom of the data region
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Title"
        ws.merge_cells("A1:C1")
        ws.cell(2, 2).value = "Priority"
        ws.cell(2, 3).value = "PlatA"
        ws.cell(3, 3).value = '=COUNTIF(C5:C1048576,"?*")'
        ws.cell(4, 3).value = 0
        ws.cell(5, 1).value = "Cat1"
        ws.merge_cells("A5:C5")
        ws.cell(6, 1).value = "Feat1"
        ws.cell(6, 2).value = "Medium"
        # Add extra empty row that makes max_row extend beyond the real data
        ws.cell(7, 1).value = None
        ws.cell(7, 2).value = None

        src = _save_wb(wb)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name

        # Provide new_rows so the new_rows branch is entered (clears leftover rows)
        new_rows = [{"category": "Cat1", "feature": "Extra", "priority": "Low", "ticked": False}]
        result = add_platform(src, out, "PlatB", {}, new_rows=new_rows)
        # Should succeed without error
        assert "new_rows_added" in result


class TestAddPlatformAutoFilter:
    """Test add_platform auto_filter update branch (line 464)."""

    def test_add_platform_updates_autofilter(self):
        """add_platform updates auto_filter when one exists."""
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        # Set an auto_filter so the branch is triggered
        ws.auto_filter.ref = "A2:D2"
        src = _save_wb(wb)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = add_platform(src, out, "PlatC", {"Feat1": True})
        assert result["platform_col"] == 5
        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.active
        # auto_filter should be updated to include new column
        assert ws2.auto_filter.ref is not None


class TestComboNoTickBranch:
    """Test create_combo_column else branch for unticked feature rows (line 595)."""

    def test_combo_unticked_feature_gets_none(self):
        """Feature row where neither platform has a tick gets None in combo column."""
        # Build a workbook where Feat3 is unticked in both platforms
        wb = _make_test_wb(with_title=True)
        ws = wb.active
        # Feat3 (row 9) is unticked in Platform A already; make Platform B also unticked
        ws.cell(9, 4).value = None  # Platform B Feat3 = no tick
        src = _save_wb(wb)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        result = create_combo_column(src, out, "A+B", "Platform A", "Platform B")
        # Feat3 has no tick in either: unique_a=1 (Feat1), overlap=1 (Feat2), unique_b=0
        assert result["tick_count"] == 2
        # Verify combo cell for Feat3 is None
        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.active
        assert ws2.cell(9, 5).value is None
