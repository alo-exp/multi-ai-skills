#!/usr/bin/env python3
"""
XLSX comparison matrix operations library.

All matrix manipulation goes through this module. The 6 Golden Rules
are enforced internally — callers never need to manage styles, merges,
or row-type detection manually.

Golden Rules (implemented in code, not left to LLM):
  1. Never hardcode styles — clone from existing cells
  2. Unmerge before writing, re-merge after
  3. Never use ws.insert_rows() — corrupts rows
  4. Row type detection: col A value + col B presence/absence
  5. Cache styles before column reorder
  6. Validate features against matrix before saving (orphan check)

Layout auto-detection:
  The module supports two XLSX layouts:
    - With title row: Row1=title(merged), Row2=headers, Row3=COUNTIF, Row4=score, Row5+=data
    - Without title row: Row1=headers, Row2=COUNTIF, Row3=score, Row4+=data
  Detection: if cell B1 == "Priority", there is no title row.

CLI usage:
    python3 matrix_ops.py add-platform --src in.xlsx --out out.xlsx \\
        --platform "Name" --features features.json [--new-rows rows.json]
    python3 matrix_ops.py reorder-columns --src in.xlsx --out out.xlsx
    python3 matrix_ops.py combo --src in.xlsx --out out.xlsx \\
        --name "A+B" --platform-a "A" --platform-b "B"
    python3 matrix_ops.py verify --src in.xlsx
    python3 matrix_ops.py reorder-rows --src in.xlsx --out out.xlsx \\
        --category "Category" --order order.json
    python3 matrix_ops.py reorder-categories --src in.xlsx --out out.xlsx \\
        --order order.json
    python3 matrix_ops.py extract-features --src in.xlsx
    python3 matrix_ops.py scores --src in.xlsx
    python3 matrix_ops.py info --src in.xlsx

All commands output JSON to stdout.
"""
from __future__ import annotations

import argparse
import copy
import json
import logging
import sys
from enum import Enum
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column constants (always the same regardless of layout)
# ---------------------------------------------------------------------------
FEAT_COL = 1      # Column A: feature name / category heading
PRIO_COL = 2      # Column B: priority string
PLAT_START = 3    # Column C: first platform

TICK = "\u2714"   # ✔

WEIGHTS: dict[str, int] = {
    "Critical": 5,
    "Very High": 4,
    "High": 3,
    "Medium": 2,
    "Low": 1,
}


class RowType(Enum):
    CATEGORY = "category"
    FEATURE = "feature"


# ---------------------------------------------------------------------------
# Layout detection
# ---------------------------------------------------------------------------

class _Layout:
    """Auto-detected matrix row layout.

    Two formats are supported:
      WITH title:    title=1, header=2, total=3, score=4, data=5
      WITHOUT title: title=None, header=1, total=2, score=3, data=4
    """
    __slots__ = ("title_row", "header_row", "total_row", "score_row", "data_start")

    def __init__(self, ws: Worksheet):
        b1 = str(ws.cell(1, PRIO_COL).value or "").strip()
        if b1 == "Priority":
            # No title row — headers are in row 1
            self.title_row: Optional[int] = None
            self.header_row = 1
            self.total_row = 2
            self.score_row = 3
            self.data_start = 4
        else:
            # Title row in row 1
            self.title_row = 1
            self.header_row = 2
            self.total_row = 3
            self.score_row = 4
            self.data_start = 5

    @property
    def freeze_pane(self) -> str:
        """Cell reference for freeze panes (freeze below header block)."""
        return f"A{self.data_start}"


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _row_type(ws: Worksheet, row: int) -> Optional[RowType]:
    """Golden Rule 4: detect row type by col-A value + col-B presence."""
    a = ws.cell(row, FEAT_COL).value
    b = ws.cell(row, PRIO_COL).value
    if not a:
        return None
    if b and str(b).strip() not in ("", "\u2014", "—"):
        return RowType.FEATURE
    return RowType.CATEGORY


def _clone_style(src, dst) -> None:
    """Golden Rule 1: clone all style attributes from one cell to another."""
    for attr in ("font", "fill", "border", "alignment", "protection"):
        v = getattr(src, attr, None)
        if v:
            setattr(dst, attr, copy.copy(v))
    dst.number_format = src.number_format


def _last_plat_col(ws: Worksheet, L: _Layout) -> int:
    """Return column index of the last platform (first empty header after PLAT_START)."""
    col = PLAT_START
    while ws.cell(L.header_row, col).value:
        col += 1
    return col - 1


def _platform_cols(ws: Worksheet, L: _Layout) -> dict[str, int]:
    """Return {platform_name: col_index} from header row."""
    result: dict[str, int] = {}
    col = PLAT_START
    while True:
        v = ws.cell(L.header_row, col).value
        if not v or not str(v).strip():  # pragma: no cover
            break
        result[str(v).strip()] = col
        col += 1
    return result


def _all_features(ws: Worksheet, L: _Layout) -> list[str]:
    """Return all feature names (not category headings) in order."""
    feats = []
    for r in range(L.data_start, ws.max_row + 1):
        if _row_type(ws, r) == RowType.FEATURE:
            feats.append(str(ws.cell(r, FEAT_COL).value))
    return feats


def _all_categories(ws: Worksheet, L: _Layout) -> list[dict]:
    """Return [{name, start_row, end_row, features: [{name, priority, row}]}]."""
    cats: list[dict] = []
    for r in range(L.data_start, ws.max_row + 1):
        rt = _row_type(ws, r)
        if rt == RowType.CATEGORY:
            cats.append({
                "name": str(ws.cell(r, FEAT_COL).value),
                "start_row": r,
                "end_row": r,
                "features": [],
            })
        elif rt == RowType.FEATURE and cats:
            cats[-1]["end_row"] = r
            cats[-1]["features"].append({
                "name": str(ws.cell(r, FEAT_COL).value),
                "priority": str(ws.cell(r, PRIO_COL).value),
                "row": r,
            })
    return cats


def _unmerge_all(ws: Worksheet) -> list[str]:
    """Golden Rule 2: unmerge every range. Returns list of range strings."""
    saved = [str(mr) for mr in ws.merged_cells.ranges]
    for mr_str in saved:
        ws.unmerge_cells(mr_str)
    return saved


def _remerge_all(ws: Worksheet, last_col: int, L: _Layout) -> None:
    """Golden Rule 2: re-merge title row (if present) and all category heading rows."""
    ltr = get_column_letter(last_col)
    # Title row (only if layout has one)
    if L.title_row is not None:
        ws.merge_cells(f"A{L.title_row}:{ltr}{L.title_row}")
    # Category heading rows
    for r in range(L.data_start, ws.max_row + 1):
        if _row_type(ws, r) == RowType.CATEGORY:
            ws.merge_cells(f"A{r}:{ltr}{r}")


def _countif(col_letter: str, L: _Layout) -> str:
    """COUNTIF formula for total capabilities row."""
    return f'=COUNTIF({col_letter}{L.data_start}:{col_letter}1048576,"?*")'


def _score_formula(col_letter: str, L: _Layout) -> str:
    """Priority-weighted COUNTIFS formula for score row."""
    b = f"$B${L.data_start}:$B$1048576"
    c = f"{col_letter}${L.data_start}:{col_letter}$1048576"
    parts = []
    for prio, wt in WEIGHTS.items():
        parts.append(f'COUNTIFS({b},"{prio}",{c},"?*")*{wt}')
    return "=" + "+".join(parts)


def _compute_score(ws: Worksheet, col: int, L: _Layout) -> int:
    """Compute weighted score for a platform column in Python."""
    score = 0
    for r in range(L.data_start, ws.max_row + 1):
        if _row_type(ws, r) != RowType.FEATURE:
            continue
        prio = str(ws.cell(r, PRIO_COL).value or "")
        tick = ws.cell(r, col).value
        if tick == TICK and prio in WEIGHTS:
            score += WEIGHTS[prio]
    return score


def _validate_no_orphans(ws: Worksheet, features: set[str], L: _Layout) -> list[str]:
    """Golden Rule 6: return orphan feature names not found in matrix."""
    matrix_feats = set(_all_features(ws, L))
    return sorted(features - matrix_feats)


def _cache_col_styles(ws: Worksheet, col: int) -> list[dict]:
    """Golden Rule 5: cache all cell styles for a column before reorder."""
    styles = []
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, col)
        styles.append({
            a: copy.copy(getattr(cell, a)) for a in
            ("font", "fill", "border", "alignment", "protection")
        })
    return styles


def _apply_col_styles(ws: Worksheet, col: int, styles: list[dict]) -> None:
    """Apply cached styles to a column."""
    for r_idx, st in enumerate(styles, start=1):
        cell = ws.cell(r_idx, col)
        for attr, val in st.items():
            setattr(cell, attr, val)


def _backfill_formulas(ws: Worksheet, last_col: int, L: _Layout) -> None:
    """Rewrite COUNTIF and score formulas for ALL platform cols."""
    for c in range(PLAT_START, last_col + 1):
        ltr = get_column_letter(c)
        ws.cell(L.total_row, c).value = _countif(ltr, L)
        ws.cell(L.score_row, c).value = _score_formula(ltr, L)


def _read_row_data(ws: Worksheet, row: int, max_col: int) -> dict:
    """Read a row's values and styles into a dict (for in-memory manipulation)."""
    values = [ws.cell(row, c).value for c in range(1, max_col + 1)]
    styles = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        styles.append({
            a: copy.copy(getattr(cell, a))
            for a in ("font", "fill", "border", "alignment", "protection")
        })
    return {"values": values, "styles": styles, "row_type": _row_type(ws, row)}


def _write_row_data(ws: Worksheet, row: int, rd: dict) -> None:
    """Write a row dict back to the worksheet."""
    for c_idx, val in enumerate(rd["values"]):
        cell = ws.cell(row, c_idx + 1)
        cell.value = val
        if c_idx < len(rd["styles"]) and rd["styles"][c_idx]:
            for attr, sval in rd["styles"][c_idx].items():
                setattr(cell, attr, copy.copy(sval))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def add_platform(
    src_xlsx: str,
    out_xlsx: str,
    platform_name: str,
    features: dict[str, bool],
    new_rows: Optional[list[dict]] = None,
) -> dict:
    """Add a new platform column to the matrix.

    Args:
        src_xlsx: Source XLSX path.
        out_xlsx: Output XLSX path.
        platform_name: Display name for the column header.
        features: {feature_name: True/False} for tick decisions.
        new_rows: Optional [{category, feature, priority, ticked}] for new rows.

    Returns:
        {ticks_applied, new_rows_added, orphans, total_features, platform_col}
    """
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb.active
    L = _Layout(ws)

    # Golden Rule 6: orphan check (only for ticked features)
    ticked_set = {f for f, v in features.items() if v}
    orphans = _validate_no_orphans(ws, ticked_set, L)

    # Golden Rule 2: unmerge before writing
    _unmerge_all(ws)

    last_col = _last_plat_col(ws, L)
    new_col = last_col + 1
    new_ltr = get_column_letter(new_col)

    # --- Insert new rows first (Golden Rule 3: no ws.insert_rows) ---
    rows_added = 0
    if new_rows:
        # Read all data into memory
        all_data = []
        for r in range(L.data_start, ws.max_row + 1):
            all_data.append(_read_row_data(ws, r, last_col))

        # Group new rows by their target category
        new_by_cat: dict[str, list[dict]] = {}
        for nr in new_rows:
            cat = nr.get("category", "")
            new_by_cat.setdefault(cat, []).append(nr)

        # Insert new rows at end of their respective categories
        expanded: list[dict] = []
        current_cat: Optional[str] = None
        pending_inserts: list[dict] = []

        for i, rd in enumerate(all_data):
            if rd["row_type"] == RowType.CATEGORY:
                # Flush pending inserts for previous category
                for nr in pending_inserts:
                    feat_name = nr["feature"]
                    prio = nr.get("priority", "Medium")
                    is_ticked = nr.get("ticked", True)
                    vals = [feat_name, prio] + [None] * (last_col - 2)
                    # Clone style from last row in expanded
                    if expanded:
                        new_rd = {
                            "values": vals,
                            "styles": [copy.copy(s) for s in expanded[-1]["styles"]],
                            "row_type": RowType.FEATURE,
                            "_new_ticked": is_ticked,
                        }
                    else:  # pragma: no cover
                        new_rd = {
                            "values": vals,
                            "styles": [{} for _ in range(last_col)],
                            "row_type": RowType.FEATURE,
                            "_new_ticked": is_ticked,
                        }
                    expanded.append(new_rd)
                    features[feat_name] = is_ticked
                    rows_added += 1

                current_cat = str(rd["values"][0]) if rd["values"][0] else ""
                pending_inserts = new_by_cat.get(current_cat, [])
                expanded.append(rd)
            else:
                expanded.append(rd)

        # Flush remaining inserts for last category
        for nr in pending_inserts:
            feat_name = nr["feature"]
            prio = nr.get("priority", "Medium")
            is_ticked = nr.get("ticked", True)
            vals = [feat_name, prio] + [None] * (last_col - 2)
            if expanded:
                new_rd = {
                    "values": vals,
                    "styles": [copy.copy(s) for s in expanded[-1]["styles"]],
                    "row_type": RowType.FEATURE,
                    "_new_ticked": is_ticked,
                }
            else:  # pragma: no cover
                new_rd = {
                    "values": vals,
                    "styles": [{} for _ in range(last_col)],
                    "row_type": RowType.FEATURE,
                    "_new_ticked": is_ticked,
                }
            expanded.append(new_rd)
            features[feat_name] = is_ticked
            rows_added += 1

        # Write expanded data back
        for idx, rd in enumerate(expanded):
            _write_row_data(ws, L.data_start + idx, rd)

        # Clear any leftover rows beyond the expanded data
        for r in range(L.data_start + len(expanded), ws.max_row + 1):  # pragma: no cover
            for c in range(1, last_col + 1):
                ws.cell(r, c).value = None

    # --- Add new platform column ---
    # Header
    hdr = ws.cell(L.header_row, new_col)
    hdr.value = platform_name
    _clone_style(ws.cell(L.header_row, last_col), hdr)

    # COUNTIF (total row) and score row
    ws.cell(L.total_row, new_col).value = _countif(new_ltr, L)
    _clone_style(ws.cell(L.total_row, last_col), ws.cell(L.total_row, new_col))

    ws.cell(L.score_row, new_col).value = _score_formula(new_ltr, L)
    _clone_style(ws.cell(L.score_row, last_col), ws.cell(L.score_row, new_col))

    # Title row: just clone style to new column (value stays merged)
    if L.title_row is not None:
        _clone_style(ws.cell(L.title_row, last_col), ws.cell(L.title_row, new_col))

    # Data rows
    ticks_applied = 0
    for r in range(L.data_start, ws.max_row + 1):
        cell = ws.cell(r, new_col)
        _clone_style(ws.cell(r, last_col), cell)
        rt = _row_type(ws, r)
        if rt == RowType.FEATURE:
            feat_name = str(ws.cell(r, FEAT_COL).value)
            if features.get(feat_name, False):
                cell.value = TICK
                ticks_applied += 1
            else:
                cell.value = None
        else:
            cell.value = None

    # Backfill formulas for ALL platform columns
    _backfill_formulas(ws, new_col, L)

    # Golden Rule 2: re-merge
    _remerge_all(ws, new_col, L)

    # Freeze panes
    ws.freeze_panes = L.freeze_pane

    # Update auto-filter if present
    if ws.auto_filter.ref:
        ws.auto_filter.ref = f"A{L.header_row}:{get_column_letter(new_col)}{L.header_row}"

    wb.save(out_xlsx)

    return {
        "platform_col": new_col,
        "ticks_applied": ticks_applied,
        "new_rows_added": rows_added,
        "orphans": orphans,
        "total_features": len(_all_features(ws, L)),
    }


def reorder_columns_by_score(src_xlsx: str, out_xlsx: str) -> dict:
    """Reorder platform columns by descending weighted score."""
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb.active
    L = _Layout(ws)

    plats = _platform_cols(ws, L)
    if not plats:
        wb.save(out_xlsx)
        return {"order": [], "scores": {}}

    # Compute scores
    scores = {name: _compute_score(ws, col, L) for name, col in plats.items()}
    sorted_names = sorted(scores, key=lambda n: -scores[n])

    # Golden Rule 2: unmerge
    _unmerge_all(ws)

    # Golden Rule 5: cache ALL styles and values before any writes
    max_row = ws.max_row
    cache: dict[str, dict] = {}
    for name, col in plats.items():
        col_data: dict[str, dict] = {"values": {}, "styles": {}}
        for r in range(1, max_row + 1):
            cell = ws.cell(r, col)
            col_data["values"][r] = cell.value
            col_data["styles"][r] = {
                a: copy.copy(getattr(cell, a))
                for a in ("font", "fill", "border", "alignment", "protection")
            }
        cache[name] = col_data

    # Write columns in new order
    for new_idx, name in enumerate(sorted_names):
        new_col = PLAT_START + new_idx
        col_data = cache[name]
        for r in range(1, max_row + 1):
            cell = ws.cell(r, new_col)
            cell.value = col_data["values"][r]
            for attr, val in col_data["styles"][r].items():
                setattr(cell, attr, copy.copy(val))

    # Rewrite formulas with correct column letters
    last_col = PLAT_START + len(sorted_names) - 1
    _backfill_formulas(ws, last_col, L)

    # Golden Rule 2: re-merge
    _remerge_all(ws, last_col, L)

    ws.freeze_panes = L.freeze_pane
    wb.save(out_xlsx)

    return {
        "order": sorted_names,
        "scores": {n: scores[n] for n in sorted_names},
    }


def create_combo_column(
    src_xlsx: str,
    out_xlsx: str,
    combo_name: str,
    platform_a: str,
    platform_b: str,
) -> dict:
    """Create a combo column = union of ticks from two platforms."""
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb.active
    L = _Layout(ws)

    plats = _platform_cols(ws, L)
    if platform_a not in plats or platform_b not in plats:
        missing = [p for p in (platform_a, platform_b) if p not in plats]
        return {"error": f"Platform(s) not found: {missing}"}

    col_a = plats[platform_a]
    col_b = plats[platform_b]

    _unmerge_all(ws)

    last_col = _last_plat_col(ws, L)
    new_col = last_col + 1
    new_ltr = get_column_letter(new_col)

    # Header
    hdr = ws.cell(L.header_row, new_col)
    hdr.value = combo_name
    _clone_style(ws.cell(L.header_row, last_col), hdr)

    # Formulas
    ws.cell(L.total_row, new_col).value = _countif(new_ltr, L)
    _clone_style(ws.cell(L.total_row, last_col), ws.cell(L.total_row, new_col))
    ws.cell(L.score_row, new_col).value = _score_formula(new_ltr, L)
    _clone_style(ws.cell(L.score_row, last_col), ws.cell(L.score_row, new_col))

    tick_count = 0
    unique_a = 0
    unique_b = 0
    overlap = 0

    for r in range(L.data_start, ws.max_row + 1):
        cell = ws.cell(r, new_col)
        _clone_style(ws.cell(r, last_col), cell)
        if _row_type(ws, r) != RowType.FEATURE:
            cell.value = None
            continue
        has_a = ws.cell(r, col_a).value == TICK
        has_b = ws.cell(r, col_b).value == TICK
        if has_a or has_b:
            cell.value = TICK
            tick_count += 1
            if has_a and has_b:
                overlap += 1
            elif has_a:
                unique_a += 1
            else:
                unique_b += 1
        else:
            cell.value = None

    _backfill_formulas(ws, new_col, L)
    _remerge_all(ws, new_col, L)
    ws.freeze_panes = L.freeze_pane
    wb.save(out_xlsx)

    return {
        "combo_col": new_col,
        "tick_count": tick_count,
        "unique_a": unique_a,
        "unique_b": unique_b,
        "overlap": overlap,
    }


def verify_ticks(src_xlsx: str) -> dict:
    """Extract per-platform tick lists for verification."""
    wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    L = _Layout(ws)

    plats = _platform_cols(ws, L)
    result: dict[str, dict] = {}

    for name, col in plats.items():
        ticked = []
        not_ticked = []
        for r in range(L.data_start, ws.max_row + 1):
            if _row_type(ws, r) != RowType.FEATURE:
                continue
            feat = str(ws.cell(r, FEAT_COL).value)
            if ws.cell(r, col).value == TICK:
                ticked.append(feat)
            else:
                not_ticked.append(feat)
        result[name] = {
            "ticked": ticked,
            "not_ticked": not_ticked,
            "tick_count": len(ticked),
            "total_features": len(ticked) + len(not_ticked),
        }

    return result


def reorder_rows(
    src_xlsx: str,
    out_xlsx: str,
    category: str,
    new_order: list[str],
) -> dict:
    """Reorder feature rows within a single category."""
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb.active
    L = _Layout(ws)

    cats = _all_categories(ws, L)
    target = None
    for cat in cats:
        if cat["name"] == category:
            target = cat
            break

    if not target:
        return {"error": f"Category not found: {category}"}

    _unmerge_all(ws)

    last_col = _last_plat_col(ws, L)

    # Read feature rows into memory (Golden Rule 3: no insert_rows)
    feat_rows: dict[str, dict] = {}
    for feat in target["features"]:
        feat_rows[feat["name"]] = _read_row_data(ws, feat["row"], last_col)

    # Write back in new order
    write_row = target["start_row"] + 1  # after category heading
    not_found = []
    reordered = 0

    # First the explicitly ordered ones
    ordered_names = []
    for name in new_order:
        if name in feat_rows:
            ordered_names.append(name)
        else:
            not_found.append(name)

    # Then remaining features not in the order list
    remaining = [f["name"] for f in target["features"]
                 if f["name"] not in set(ordered_names)]
    final_order = ordered_names + remaining

    for name in final_order:
        _write_row_data(ws, write_row, feat_rows[name])
        write_row += 1
        reordered += 1

    _remerge_all(ws, last_col, L)
    ws.freeze_panes = L.freeze_pane
    wb.save(out_xlsx)

    return {
        "category": category,
        "rows_reordered": reordered,
        "features_not_found": not_found,
    }


def reorder_categories(
    src_xlsx: str,
    out_xlsx: str,
    new_order: list[str],
) -> dict:
    """Reorder category blocks in the matrix."""
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb.active
    L = _Layout(ws)

    cats = _all_categories(ws, L)
    last_col = _last_plat_col(ws, L)

    _unmerge_all(ws)

    # Read all category blocks into memory
    blocks: dict[str, list[dict]] = {}
    for cat in cats:
        rows = []
        for r in range(cat["start_row"], cat["end_row"] + 1):
            rows.append(_read_row_data(ws, r, last_col))
        blocks[cat["name"]] = rows

    # Build final order: explicit order first, then remaining
    not_found = []
    ordered = []
    for name in new_order:
        if name in blocks:
            ordered.append(name)
        else:
            not_found.append(name)
    remaining = [c["name"] for c in cats if c["name"] not in set(ordered)]
    final = ordered + remaining

    # Write back
    write_row = L.data_start
    for cat_name in final:
        for rd in blocks[cat_name]:
            _write_row_data(ws, write_row, rd)
            write_row += 1

    _remerge_all(ws, last_col, L)
    ws.freeze_panes = L.freeze_pane
    wb.save(out_xlsx)

    return {
        "categories_reordered": len(ordered),
        "categories_not_found": not_found,
    }


def extract_features(src_xlsx: str) -> dict:
    """Extract all features grouped by category."""
    wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    L = _Layout(ws)

    cats = _all_categories(ws, L)
    result = []
    total = 0
    for cat in cats:
        feats = [{"name": f["name"], "priority": f["priority"]}
                 for f in cat["features"]]
        result.append({"name": cat["name"], "features": feats})
        total += len(feats)

    return {
        "categories": result,
        "total_categories": len(result),
        "total_features": total,
    }


def ranked_scores(src_xlsx: str) -> dict:
    """Return ranked platform scores."""
    wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    L = _Layout(ws)

    plats = _platform_cols(ws, L)
    rankings = []
    for name, col in plats.items():
        score = _compute_score(ws, col, L)
        ticks = sum(
            1 for r in range(L.data_start, ws.max_row + 1)
            if _row_type(ws, r) == RowType.FEATURE and ws.cell(r, col).value == TICK
        )
        rankings.append({"platform": name, "score": score, "ticks": ticks})

    rankings.sort(key=lambda x: -x["score"])
    for i, r in enumerate(rankings):
        r["rank"] = i + 1

    return {"rankings": rankings}


def info(src_xlsx: str) -> dict:
    """Return summary info about the matrix."""
    wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    L = _Layout(ws)

    plats = _platform_cols(ws, L)
    cats = _all_categories(ws, L)
    total_feats = sum(len(c["features"]) for c in cats)
    title = str(ws.cell(1, 1).value or "")

    return {
        "title": title,
        "platforms": list(plats.keys()),
        "platform_count": len(plats),
        "categories": [c["name"] for c in cats],
        "category_count": len(cats),
        "total_features": total_feats,
        "layout": "with_title" if L.title_row else "no_title",
        "header_row": L.header_row,
        "data_start_row": L.data_start,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _json_out(data: Any) -> None:
    """Print JSON to stdout."""
    print(json.dumps(data, indent=2, ensure_ascii=False))


def main():
    parser = argparse.ArgumentParser(
        description="XLSX comparison matrix operations",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # add-platform
    p = sub.add_parser("add-platform", help="Add a new platform column")
    p.add_argument("--src", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--platform", required=True, help="Platform display name")
    p.add_argument("--features", required=True, help="JSON file: {feat: true/false}")
    p.add_argument("--new-rows", default=None,
                   help="JSON file: [{category, feature, priority, ticked}]")

    # reorder-columns
    p = sub.add_parser("reorder-columns", help="Reorder columns by score")
    p.add_argument("--src", required=True)
    p.add_argument("--out", required=True)

    # combo
    p = sub.add_parser("combo", help="Create combo column (union of two)")
    p.add_argument("--src", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--name", required=True, help="Combo column name")
    p.add_argument("--platform-a", required=True)
    p.add_argument("--platform-b", required=True)

    # verify
    p = sub.add_parser("verify", help="Extract per-platform tick lists")
    p.add_argument("--src", required=True)

    # reorder-rows
    p = sub.add_parser("reorder-rows", help="Reorder features within a category")
    p.add_argument("--src", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--category", required=True)
    p.add_argument("--order", required=True, help="JSON file: [feat1, feat2, ...]")

    # reorder-categories
    p = sub.add_parser("reorder-categories", help="Reorder category blocks")
    p.add_argument("--src", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--order", required=True, help="JSON file: [cat1, cat2, ...]")

    # extract-features
    p = sub.add_parser("extract-features", help="List all features by category")
    p.add_argument("--src", required=True)

    # scores
    p = sub.add_parser("scores", help="Ranked platform scores")
    p.add_argument("--src", required=True)

    # info
    p = sub.add_parser("info", help="Matrix summary info")
    p.add_argument("--src", required=True)

    args = parser.parse_args()

    if args.command == "add-platform":
        feats = json.loads(Path(args.features).read_text(encoding="utf-8"))
        new_rows = None
        if args.new_rows:
            new_rows = json.loads(Path(args.new_rows).read_text(encoding="utf-8"))
        _json_out(add_platform(args.src, args.out, args.platform, feats, new_rows))

    elif args.command == "reorder-columns":
        _json_out(reorder_columns_by_score(args.src, args.out))

    elif args.command == "combo":
        _json_out(create_combo_column(
            args.src, args.out, args.name, args.platform_a, args.platform_b))

    elif args.command == "verify":
        _json_out(verify_ticks(args.src))

    elif args.command == "reorder-rows":
        order = json.loads(Path(args.order).read_text(encoding="utf-8"))
        _json_out(reorder_rows(args.src, args.out, args.category, order))

    elif args.command == "reorder-categories":
        order = json.loads(Path(args.order).read_text(encoding="utf-8"))
        _json_out(reorder_categories(args.src, args.out, order))

    elif args.command == "extract-features":
        _json_out(extract_features(args.src))

    elif args.command == "scores":
        _json_out(ranked_scores(args.src))

    elif args.command == "info":
        _json_out(info(args.src))


if __name__ == "__main__":
    main()
